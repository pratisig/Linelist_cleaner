"""
Data Quality Profiling, Auditing, and Multi-Sheet Excel/HTML Report Generator with Spatial P-Code Support - V2.
PratiSIG Consulting Services - Dakar, Sénégal.
Auteur : Youssoupha MBODJI (pratisig.consulting@gmail.com)
Version: 2.0.0 - Workbook now 6 tabs + V2 KPIs
"""

import io
from typing import Dict, List, Optional, Tuple, Any
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from linelist_cleaner.schemas.models import (
    ColumnProfile,
    DataQualityScores,
    ValidationIssue,
    DuplicateGroup,
    CleaningReport,
    SpatialCascadeSummary,
)
from linelist_cleaner.schemas.epi_dictionary import CANONICAL_TAGS


class DataQualityAuditor:
    """Calculates comprehensive Linelist Data Quality Index and generates audit reports."""

    @staticmethod
    def calculate_quality_scores(
        df: pd.DataFrame,
        issues: List[ValidationIssue],
        duplicates: List[DuplicateGroup],
        tag_to_col: Dict[str, str]
    ) -> DataQualityScores:
        n_rows = len(df)
        if n_rows == 0:
            return DataQualityScores(
                overall_score=100.0,
                grade="A",
                completeness_score=100.0,
                chronology_score=100.0,
                validity_score=100.0,
                uniqueness_score=100.0
            )
        key_tags = ["case_id", "sex", "age", "date_onset", "outcome", "case_definition"]
        key_cols = [tag_to_col[t] for t in key_tags if t in tag_to_col and tag_to_col[t] in df.columns]
        if not key_cols:
            key_cols = df.columns.tolist()
        completeness_pcts = []
        for col in key_cols:
            valid_pct = (df[col].notna() & (df[col].astype(str).str.strip() != "")).mean() * 100
            completeness_pcts.append(valid_pct)
        completeness_score = float(np.mean(completeness_pcts)) if completeness_pcts else 100.0
        chrono_errors = sum(1 for i in issues if i.issue_type == "DATE_CHRONOLOGY" and i.severity == "ERROR")
        chrono_score = max(0.0, 100.0 - (chrono_errors / n_rows * 100.0 * 2.0))
        other_errors = sum(1 for i in issues if i.issue_type != "DATE_CHRONOLOGY" and i.severity == "ERROR")
        warnings = sum(1 for i in issues if i.severity == "WARNING")
        validity_penalty = (other_errors * 2.0 + warnings * 0.5) / n_rows * 100.0
        validity_score = max(0.0, 100.0 - validity_penalty)
        dup_row_count = sum(len(g.row_indices) - 1 for g in duplicates)
        uniqueness_score = max(0.0, 100.0 - (dup_row_count / n_rows * 100.0 * 2.5))
        overall = (
            0.30 * completeness_score +
            0.25 * chrono_score +
            0.25 * validity_score +
            0.20 * uniqueness_score
        )
        overall = round(max(0.0, min(100.0, overall)), 1)
        if overall >= 90:
            grade = "A"
        elif overall >= 80:
            grade = "B"
        elif overall >= 70:
            grade = "C"
        elif overall >= 60:
            grade = "D"
        else:
            grade = "F"
        return DataQualityScores(
            overall_score=overall,
            grade=grade,
            completeness_score=round(completeness_score, 1),
            chronology_score=round(chrono_score, 1),
            validity_score=round(validity_score, 1),
            uniqueness_score=round(uniqueness_score, 1)
        )

    @staticmethod
    def profile_columns(
        df: pd.DataFrame,
        tag_to_col: Dict[str, str],
        issues: List[ValidationIssue]
    ) -> Dict[str, ColumnProfile]:
        profiles: Dict[str, ColumnProfile] = {}
        col_to_tag = {v: k for k, v in tag_to_col.items()}
        issue_counts_by_col: Dict[str, int] = {}
        for issue in issues:
            if issue.column:
                issue_counts_by_col[issue.column] = issue_counts_by_col.get(issue.column, 0) + 1
        n_rows = len(df)
        for col in df.columns:
            series = df[col]
            missing_count = int((series.isna() | (series.astype(str).str.strip() == "")).sum())
            missing_pct = round((missing_count / n_rows) * 100, 1) if n_rows > 0 else 0.0
            valid_series = series[series.notna() & (series.astype(str).str.strip() != "")]
            unique_count = int(valid_series.nunique())
            top_vals = {}
            if not valid_series.empty:
                vc = valid_series.astype(str).value_counts().head(5)
                top_vals = {str(k): int(v) for k, v in vc.items()}
            samples = [str(x) for x in valid_series.head(4).tolist()]
            inferred = "string"
            if pd.api.types.is_numeric_dtype(series):
                inferred = "numeric"
            elif pd.api.types.is_datetime64_any_dtype(series):
                inferred = "date"
            elif col_to_tag.get(col) in CANONICAL_TAGS:
                inferred = CANONICAL_TAGS[col_to_tag[col]]["type"]
            profiles[col] = ColumnProfile(
                column_name=col,
                inferred_type=inferred,
                mapped_tag=col_to_tag.get(col),
                total_count=n_rows,
                missing_count=missing_count,
                missing_percentage=missing_pct,
                unique_count=unique_count,
                top_values=top_vals,
                sample_values=samples,
                issue_count=issue_counts_by_col.get(col, 0)
            )
        return profiles

    @staticmethod
    def export_excel_audit_workbook(
        df_clean: pd.DataFrame,
        report: CleaningReport,
        output_path_or_buffer: Any,
        ref_df: Optional[pd.DataFrame] = None
    ) -> None:
        """
        Exports V2 6-tab Workbook:
        - Tab 1: KPI_Dashboard (V2 with coords, phones, outbreak alerts, trend)
        - Tab 2: LineList_Nettoyee (colored by MATCH_LEVEL)
        - Tab 3: Referentiel_PCode
        - Tab 4: Anomalies_Qualite (validation issues)
        - Tab 5: Doublons (duplicate groups)
        - Tab 6: Dictionnaire_Donnees (column profiles + tags)
        """
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        navy_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        blue_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        teal_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
        slate_fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
        amber_fill = PatternFill(start_color="D97706", end_color="D97706", fill_type="solid")
        emerald_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
        rose_fill = PatternFill(start_color="E11D48", end_color="E11D48", fill_type="solid")

        match_colors = {
            "Locality": PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"),
            "Admin3_Ward": PatternFill(start_color="CCFBF1", end_color="CCFBF1", fill_type="solid"),
            "Admin2_LGA": PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
            "Admin1_State": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
            "Unmatched": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
        }

        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        title_font = Font(name="Calibri", size=14, bold=True, color="0F172A")
        sub_title_font = Font(name="Calibri", size=10, italic=True, color="475569")
        section_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        bold_font = Font(name="Calibri", size=10, bold=True)
        regular_font = Font(name="Calibri", size=10)
        thin_border = Border(
            left=Side(style="thin", color="E2E8F0"),
            right=Side(style="thin", color="E2E8F0"),
            top=Side(style="thin", color="E2E8F0"),
            bottom=Side(style="thin", color="E2E8F0"),
        )

        # -------------------------------------------------------------
        # TAB 1: KPI_Dashboard V2
        # -------------------------------------------------------------
        ws_kpi = wb.create_sheet(title="KPI_Dashboard")
        ws_kpi.views.sheetView[0].showGridLines = True
        ws_kpi.sheet_properties.pageSetUpPr.fitToPage = True

        ws_kpi["A1"] = "TABLEAU DE BORD V2 — QUALITE, GEOCODAGE & VEILLE EPIDEMIOLOGIQUE"
        ws_kpi["A1"].font = title_font
        ws_kpi["A2"] = "PratiSIG Consulting Services - Dakar, Sénégal | V2.0.0 | Auteur : Youssoupha MBODJI (pratisig.consulting@gmail.com) | La pratique des SIG, notre métier"
        ws_kpi["A2"].font = sub_title_font
        ws_kpi.merge_cells("A1:F1"); ws_kpi.merge_cells("A2:F2")

        spatial = report.spatial_summary
        total_rows = report.cleaned_shape[0]
        if spatial:
            geo_rate = f"{spatial.geocoded_rate_pct}%"
            geo_count = f"{spatial.geocoded_count} / {spatial.total_records}"
            avg_score = f"{spatial.average_match_score}%"
        else:
            geo_rate = "N/A"
            geo_count = f"{total_rows} records"
            avg_score = "N/A"

        # Trend & alerts
        trend = getattr(report, "incidence_trend", None)
        alerts = getattr(report, "outbreak_alerts", []) or []
        trend_label = f"{trend.trend} ({trend.weekly_growth_pct}%)" if trend and hasattr(trend, "trend") else (trend.get("trend") + f" ({trend.get('weekly_growth_pct')}%)" if isinstance(trend, dict) and trend else "N/A")
        if trend and isinstance(trend, dict):
            if trend.get("trend") == "increasing":
                trend_label = f"↗ Hausse ({trend.get('weekly_growth_pct')}%)"
            elif trend.get("trend") == "decreasing":
                trend_label = f"↘ Baisse ({trend.get('weekly_growth_pct')}%)"
        elif trend:
            try:
                t = trend.trend
                g = trend.weekly_growth_pct
                trend_label = f"{'↗ Hausse' if t=='increasing' else '↘ Baisse' if t=='decreasing' else '→ Stable'} ({g}%)"
            except:
                pass

        # V2 metrics
        coords_cleaned = getattr(report, "coordinates_cleaned", 0)
        phones_std = getattr(report, "phones_standardized", 0)

        kpi_metrics = [
            ("Nombre Total de Cas", str(total_rows), "Lignes exploitables après validation"),
            ("Taux Global de Géocodage", geo_rate, "Proportion rattachée à un P-Code valide"),
            ("Cas Géocodés", geo_count, "Nombre absolu localisés"),
            ("Score Moyen Similarité Fuzzy", avg_score, "Précision moyenne du rapprochement"),
            ("Score Global Qualité", f"{report.quality_scores_after.overall_score}%", f"Grade: {report.quality_scores_after.grade}"),
            ("Semaines Epi Calculées", f"{report.epi_weeks_computed} / {total_rows}", "Dates normalisées vers EPI_WEEK OMS"),
            ("Coordonnées Validées (V2)", str(coords_cleaned), "Lat/Lon WGS84 nettoyées / swaps corrigés"),
            ("Téléphones Normalisés (V2)", str(phones_std), "Format international E.164"),
            ("Tendance Incidence (V2)", trend_label, f"Semaine pic: {getattr(trend,'peak_week', None) or (trend.get('peak_week') if isinstance(trend, dict) else 'N/A')}"),
            ("Alertes Épidémiques (V2)", str(len(alerts)), "Semaines > seuil anomalie"),
        ]

        ws_kpi["A4"] = "1. Indicateurs Clés — Géocodage & Qualité V2"
        ws_kpi["A4"].font = section_font; ws_kpi["A4"].fill = blue_fill; ws_kpi.merge_cells("A4:F4")
        for idx, (label, val, note) in enumerate(kpi_metrics, start=5):
            ws_kpi[f"A{idx}"] = label; ws_kpi[f"A{idx}"].font = bold_font
            ws_kpi[f"B{idx}"] = val; ws_kpi[f"B{idx}"].font = bold_font; ws_kpi[f"B{idx}"].alignment = Alignment(horizontal="center")
            ws_kpi[f"C{idx}"] = note; ws_kpi[f"C{idx}"].font = regular_font
            for c in ["A","B","C","D","E","F"]:
                ws_kpi[f"{c}{idx}"].border = thin_border
            # highlight V2 rows
            if "V2" in label:
                ws_kpi[f"A{idx}"].fill = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid")

        # Cascade breakdown
        start_cascade_row = len(kpi_metrics) + 7
        ws_kpi[f"A{start_cascade_row}"] = "2. Répartition par Niveau de Résolution Spatiale (Cascade)"
        ws_kpi[f"A{start_cascade_row}"].font = section_font; ws_kpi[f"A{start_cascade_row}"].fill = teal_fill; ws_kpi.merge_cells(f"A{start_cascade_row}:F{start_cascade_row}")
        casc_headers = ["Niveau (MATCH_LEVEL)", "Description", "Cas", "Proportion (%)", "Couleur"]
        h_row = start_cascade_row + 1
        for col_i, h_text in enumerate(casc_headers, start=1):
            cell = ws_kpi.cell(row=h_row, column=col_i, value=h_text); cell.font = header_font; cell.fill = navy_fill
        levels_info = [
            ("Locality", "Village / Localité / Camp IDP (Précision max)", match_colors["Locality"]),
            ("Admin3_Ward", "Admin 3 / Ward / Aire de santé (Fallback 1)", match_colors["Admin3_Ward"]),
            ("Admin2_LGA", "Admin 2 / LGA / Zone de santé (Fallback 2)", match_colors["Admin2_LGA"]),
            ("Admin1_State", "Admin 1 / État / Province (Fallback 3)", match_colors["Admin1_State"]),
            ("Unmatched", "Non localisé", match_colors["Unmatched"])
        ]
        curr_r = h_row + 1
        for lvl, desc, fill_style in levels_info:
            cnt = spatial.level_distribution.get(lvl, 0) if spatial else 0
            pct = spatial.level_percentages.get(lvl, 0.0) if spatial else 0.0
            ws_kpi.cell(row=curr_r, column=1, value=lvl).font = bold_font
            ws_kpi.cell(row=curr_r, column=2, value=desc).font = regular_font
            ws_kpi.cell(row=curr_r, column=3, value=cnt).font = bold_font
            ws_kpi.cell(row=curr_r, column=4, value=f"{pct}%").font = bold_font
            c_cell = ws_kpi.cell(row=curr_r, column=5, value="●"); c_cell.font = bold_font; c_cell.fill = fill_style; c_cell.alignment = Alignment(horizontal="center")
            for c in range(1, 6):
                ws_kpi.cell(row=curr_r, column=c).border = thin_border
            curr_r += 1

        # EpiWeek Distribution
        if "EPI_WEEK" in df_clean.columns:
            start_epi_row = curr_r + 2
            ws_kpi[f"A{start_epi_row}"] = "3. Synthèse par Semaine Épidémiologique OMS (EPI_WEEK)"
            ws_kpi[f"A{start_epi_row}"].font = section_font; ws_kpi[f"A{start_epi_row}"].fill = blue_fill; ws_kpi.merge_cells(f"A{start_epi_row}:F{start_epi_row}")
            epi_h_row = start_epi_row + 1
            for col_i, h_text in enumerate(["Semaine Epi", "Cas", "Proportion (%)"], start=1):
                cell = ws_kpi.cell(row=epi_h_row, column=col_i, value=h_text); cell.font = header_font; cell.fill = navy_fill
            epi_counts = df_clean["EPI_WEEK"].dropna().value_counts().sort_index()
            curr_epi_r = epi_h_row + 1
            for wk, c_num in epi_counts.items():
                p_val = round((c_num / total_rows) * 100, 1) if total_rows > 0 else 0.0
                ws_kpi.cell(row=curr_epi_r, column=1, value=str(wk)).font = bold_font
                ws_kpi.cell(row=curr_epi_r, column=2, value=int(c_num)).font = regular_font
                ws_kpi.cell(row=curr_epi_r, column=3, value=f"{p_val}%").font = regular_font
                for c in range(1, 4):
                    ws_kpi.cell(row=curr_epi_r, column=c).border = thin_border
                curr_epi_r += 1
            # Alerts section after epi
            if alerts:
                alert_start = curr_epi_r + 2
                ws_kpi[f"A{alert_start}"] = "4. Alertes Épidémiques Détectées (V2 Anomaly Detection)"
                ws_kpi[f"A{alert_start}"].font = section_font; ws_kpi[f"A{alert_start}"].fill = rose_fill; ws_kpi.merge_cells(f"A{alert_start}:F{alert_start}")
                hdr_row = alert_start + 1
                for ci, ht in enumerate(["Semaine", "Cas", "Moy. baseline", "Sévérité", "Message"], start=1):
                    cell = ws_kpi.cell(row=hdr_row, column=ci, value=ht); cell.font = header_font; cell.fill = navy_fill
                r = hdr_row + 1
                for al in alerts[:10]:
                    # handle both dict and model
                    if isinstance(al, dict):
                        wk = al.get("epi_week"); cases = al.get("cases"); base = al.get("baseline_mean"); sev = al.get("severity"); msg = al.get("message")
                    else:
                        wk = al.epi_week; cases = al.cases; base = al.baseline_mean; sev = al.severity; msg = al.message
                    ws_kpi.cell(row=r, column=1, value=str(wk)).font = bold_font
                    ws_kpi.cell(row=r, column=2, value=int(cases)).font = bold_font
                    ws_kpi.cell(row=r, column=3, value=float(base)).font = regular_font
                    sev_cell = ws_kpi.cell(row=r, column=4, value=str(sev)); sev_cell.font = bold_font
                    sev_cell.fill = rose_fill if sev=="HIGH" else amber_fill
                    ws_kpi.cell(row=r, column=5, value=str(msg)).font = regular_font
                    for c in range(1,6):
                        ws_kpi.cell(row=r, column=c).border = thin_border
                    r += 1
                curr_epi_r = r
            footer_r = curr_epi_r + 2
        else:
            footer_r = curr_r + 6

        ws_kpi[f"A{footer_r}"] = "© PratiSIG Consulting Services — V2.0.0 — Outil libre communauté SIG & humanitaire | Export GeoJSON disponible via API /export/geojson"
        ws_kpi[f"A{footer_r}"].font = sub_title_font; ws_kpi.merge_cells(f"A{footer_r}:F{footer_r}")
        ws_kpi.column_dimensions["A"].width = 32; ws_kpi.column_dimensions["B"].width = 22; ws_kpi.column_dimensions["C"].width = 40; ws_kpi.column_dimensions["D"].width = 16; ws_kpi.column_dimensions["E"].width = 16; ws_kpi.column_dimensions["F"].width = 18

        # -------------------------------------------------------------
        # TAB 2: LineList_Nettoyee
        # -------------------------------------------------------------
        ws_data = wb.create_sheet(title="LineList_Nettoyee")
        ws_data.views.sheetView[0].showGridLines = True
        headers = list(df_clean.columns)
        for col_idx, col_name in enumerate(headers, start=1):
            cell = ws_data.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font; cell.fill = navy_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        match_level_col_idx = headers.index("MATCH_LEVEL") if "MATCH_LEVEL" in headers else None
        for row_idx, row_values in enumerate(df_clean.itertuples(index=False), start=2):
            match_lvl = row_values[match_level_col_idx] if match_level_col_idx is not None else None
            lvl_fill = match_colors.get(str(match_lvl), None)
            for col_idx, val in enumerate(row_values, start=1):
                val_to_write = "" if pd.isna(val) else val
                cell = ws_data.cell(row=row_idx, column=col_idx, value=val_to_write)
                cell.font = regular_font; cell.border = thin_border
                if col_idx - 1 == match_level_col_idx and lvl_fill:
                    cell.fill = lvl_fill; cell.font = bold_font
                elif val_to_write == "":
                    cell.fill = PatternFill(start_color="F8FAFC", fill_type="solid")
        for col in ws_data.columns:
            max_len = max(len(str(cell.value or "")) for cell in col[:40])
            col_letter = get_column_letter(col[0].column)
            ws_data.column_dimensions[col_letter].width = max(min(max_len + 3, 32), 12)
        ws_data.freeze_panes = "A2"
        ws_data.auto_filter.ref = ws_data.dimensions

        # -------------------------------------------------------------
        # TAB 3: Referentiel_PCode
        # -------------------------------------------------------------
        ws_ref = wb.create_sheet(title="Referentiel_PCode")
        ws_ref.views.sheetView[0].showGridLines = True
        if ref_df is not None and not ref_df.empty:
            ref_headers = list(ref_df.columns)
            for c_idx, h_text in enumerate(ref_headers, start=1):
                cell = ws_ref.cell(row=1, column=c_idx, value=h_text); cell.font = header_font; cell.fill = teal_fill; cell.alignment = Alignment(horizontal="center", vertical="center")
            for r_idx, r_values in enumerate(ref_df.itertuples(index=False), start=2):
                for c_idx, val in enumerate(r_values, start=1):
                    cell = ws_ref.cell(row=r_idx, column=c_idx, value="" if pd.isna(val) else val); cell.font = regular_font; cell.border = thin_border
            for col in ws_ref.columns:
                max_len = max(len(str(cell.value or "")) for cell in col[:30])
                col_letter = get_column_letter(col[0].column); ws_ref.column_dimensions[col_letter].width = max(max_len + 3, 14)
            ws_ref.freeze_panes = "A2"; ws_ref.auto_filter.ref = ws_ref.dimensions
        else:
            ws_ref["A1"] = "Aucun référentiel P-Code externe chargé — utilisation du référentiel intégré si disponible."
            ws_ref["A1"].font = regular_font; ws_ref["A2"] = "Astuce V2: Chargez votre propre référentiel COD-AB via la zone verte dans l'interface web."; ws_ref["A2"].font = sub_title_font

        # -------------------------------------------------------------
        # TAB 4: Anomalies_Qualite (Validation Issues)
        # -------------------------------------------------------------
        ws_iss = wb.create_sheet(title="Anomalies_Qualite")
        ws_iss.views.sheetView[0].showGridLines = True
        iss_headers = ["Ligne", "Case ID", "Sévérité", "Type", "Colonne", "Message", "Action suggérée", "Valeur brute"]
        for ci, h in enumerate(iss_headers, start=1):
            cell = ws_iss.cell(row=1, column=ci, value=h); cell.font = header_font; cell.fill = rose_fill; cell.alignment = Alignment(horizontal="center")
        issues = report.validation_issues or []
        # Sort by severity ERROR first
        sev_order = {"ERROR":0, "WARNING":1, "INFO":2}
        def sev_key(x):
            if isinstance(x, dict):
                return sev_order.get(x.get("severity","INFO"), 2)
            return sev_order.get(getattr(x,"severity","INFO"), 2)
        issues_sorted = sorted(issues, key=sev_key)
        for r, iss in enumerate(issues_sorted[:2000], start=2):  # cap 2000 rows
            if isinstance(iss, dict):
                row_idx = iss.get("row_idx"); case_id = iss.get("case_id"); sev = iss.get("severity"); itype = iss.get("issue_type"); col = iss.get("column"); msg = iss.get("message"); sugg = iss.get("suggested_action"); raw = iss.get("raw_value")
            else:
                row_idx = iss.row_idx; case_id = iss.case_id; sev = iss.severity; itype = iss.issue_type; col = iss.column; msg = iss.message; sugg = iss.suggested_action; raw = iss.raw_value
            ws_iss.cell(row=r, column=1, value=row_idx).font = regular_font
            ws_iss.cell(row=r, column=2, value=str(case_id) if case_id else "").font = regular_font
            sev_cell = ws_iss.cell(row=r, column=3, value=str(sev)); sev_cell.font = bold_font; sev_cell.alignment = Alignment(horizontal="center")
            if sev=="ERROR": sev_cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
            elif sev=="WARNING": sev_cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
            else: sev_cell.fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
            ws_iss.cell(row=r, column=4, value=str(itype)).font = regular_font
            ws_iss.cell(row=r, column=5, value=str(col) if col else "").font = regular_font
            ws_iss.cell(row=r, column=6, value=str(msg)).font = regular_font; ws_iss.cell(row=r, column=6).alignment = Alignment(wrap_text=True)
            ws_iss.cell(row=r, column=7, value=str(sugg) if sugg else "").font = regular_font; ws_iss.cell(row=r, column=7).alignment = Alignment(wrap_text=True)
            ws_iss.cell(row=r, column=8, value=str(raw) if raw is not None else "").font = regular_font
            for c in range(1,9):
                ws_iss.cell(row=r, column=c).border = thin_border
        if not issues:
            ws_iss["A2"] = "Aucune anomalie détectée — félicitations ! Qualité élevée."; ws_iss["A2"].font = bold_font
        ws_iss.freeze_panes = "A2"; ws_iss.auto_filter.ref = ws_iss.dimensions
        for idx,w in enumerate([8,14,12,20,16,44,30,18], start=1):
            ws_iss.column_dimensions[get_column_letter(idx)].width = w

        # -------------------------------------------------------------
        # TAB 5: Doublons
        # -------------------------------------------------------------
        ws_dup = wb.create_sheet(title="Doublons")
        ws_dup.views.sheetView[0].showGridLines = True
        dup_headers = ["Groupe #", "Type", "Score", "Indices lignes", "Case IDs", "Idx recommandé à garder", "Clés de matching"]
        for ci, h in enumerate(dup_headers, start=1):
            cell = ws_dup.cell(row=1, column=ci, value=h); cell.font = header_font; cell.fill = amber_fill; cell.alignment = Alignment(horizontal="center")
        dup_groups = report.duplicate_groups or []
        for r, g in enumerate(dup_groups[:500], start=2):
            if isinstance(g, dict):
                gid = g.get("group_id"); dtype = g.get("duplicate_type"); score = g.get("match_score"); idxs = g.get("row_indices"); cids = g.get("case_ids"); rec = g.get("recommended_keep_idx"); keys = g.get("matching_keys")
            else:
                gid = g.group_id; dtype = g.duplicate_type; score = g.match_score; idxs = g.row_indices; cids = g.case_ids; rec = g.recommended_keep_idx; keys = g.matching_keys
            ws_dup.cell(row=r, column=1, value=int(gid) if gid is not None else r-1).font = bold_font
            ws_dup.cell(row=r, column=2, value=str(dtype)).font = regular_font
            ws_dup.cell(row=r, column=3, value=float(score) if score is not None else 0).font = regular_font; ws_dup.cell(row=r, column=3).number_format = "0.00"
            ws_dup.cell(row=r, column=4, value=str(idxs)).font = regular_font
            ws_dup.cell(row=r, column=5, value=str(cids)).font = regular_font
            ws_dup.cell(row=r, column=6, value=int(rec) if rec is not None else "").font = bold_font
            ws_dup.cell(row=r, column=7, value=str(keys)).font = regular_font; ws_dup.cell(row=r, column=7).alignment = Alignment(wrap_text=True)
            for c in range(1,8):
                ws_dup.cell(row=r, column=c).border = thin_border
        if not dup_groups:
            ws_dup["A2"] = "Aucun doublon détecté."; ws_dup["A2"].font = bold_font
        ws_dup.freeze_panes = "A2"; ws_dup.auto_filter.ref = ws_dup.dimensions
        for idx,w in enumerate([10,12,10,18,28,18,36], start=1):
            ws_dup.column_dimensions[get_column_letter(idx)].width = w

        # -------------------------------------------------------------
        # TAB 6: Dictionnaire_Donnees + Profil colonnes
        # -------------------------------------------------------------
        ws_dict = wb.create_sheet(title="Dictionnaire_Donnees")
        ws_dict.views.sheetView[0].showGridLines = True
        dict_headers = ["Colonne", "Tag Epi", "Type inféré", "Total", "Manquants", "% Manquant", "Uniques", "Top valeurs", "Exemples", "Issues"]
        for ci, h in enumerate(dict_headers, start=1):
            cell = ws_dict.cell(row=1, column=ci, value=h); cell.font = header_font; cell.fill = slate_fill; cell.alignment = Alignment(horizontal="center", wrap_text=True)
        profiles = report.column_profiles or {}
        # profiles may be dict of ColumnProfile
        r = 2
        for col_name, prof in profiles.items():
            if isinstance(prof, dict):
                tag = prof.get("mapped_tag"); itype = prof.get("inferred_type"); total = prof.get("total_count"); miss = prof.get("missing_count"); miss_pct = prof.get("missing_percentage"); uniq = prof.get("unique_count"); top = prof.get("top_values"); samples = prof.get("sample_values"); iss_cnt = prof.get("issue_count",0)
            else:
                tag = prof.mapped_tag; itype = prof.inferred_type; total = prof.total_count; miss = prof.missing_count; miss_pct = prof.missing_percentage; uniq = prof.unique_count; top = prof.top_values; samples = prof.sample_values; iss_cnt = prof.issue_count
            ws_dict.cell(row=r, column=1, value=str(col_name)).font = bold_font
            ws_dict.cell(row=r, column=2, value=str(tag) if tag else "—").font = regular_font
            ws_dict.cell(row=r, column=3, value=str(itype)).font = regular_font
            ws_dict.cell(row=r, column=4, value=int(total) if total is not None else 0).font = regular_font
            ws_dict.cell(row=r, column=5, value=int(miss) if miss is not None else 0).font = regular_font
            ws_dict.cell(row=r, column=6, value=float(miss_pct) if miss_pct is not None else 0).font = regular_font; ws_dict.cell(row=r, column=6).number_format = "0.0"
            ws_dict.cell(row=r, column=7, value=int(uniq) if uniq is not None else 0).font = regular_font
            ws_dict.cell(row=r, column=8, value=str(top) if top else "").font = regular_font; ws_dict.cell(row=r, column=8).alignment = Alignment(wrap_text=True)
            ws_dict.cell(row=r, column=9, value=str(samples) if samples else "").font = regular_font; ws_dict.cell(row=r, column=9).alignment = Alignment(wrap_text=True)
            iss_cell = ws_dict.cell(row=r, column=10, value=int(iss_cnt) if iss_cnt else 0); iss_cell.font = bold_font; iss_cell.alignment = Alignment(horizontal="center")
            if iss_cnt and iss_cnt > 0:
                iss_cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
            for c in range(1,11):
                ws_dict.cell(row=r, column=c).border = thin_border
            r += 1
        # Add footer dictionary legend
        leg_r = r + 2
        ws_dict[f"A{leg_r}"] = "Légende V2 — Tags canoniques & catégories"
        ws_dict[f"A{leg_r}"].font = section_font; ws_dict[f"A{leg_r}"].fill = slate_fill; ws_dict.merge_cells(f"A{leg_r}:J{leg_r}")
        # list canonical tags categories
        cat_rows = {}
        for tag, meta in CANONICAL_TAGS.items():
            cat = meta.get("category","other")
            cat_rows.setdefault(cat, []).append(tag)
        lr = leg_r + 1
        for cat, tags in cat_rows.items():
            ws_dict.cell(row=lr, column=1, value=cat).font = bold_font
            ws_dict.cell(row=lr, column=2, value=", ".join(tags[:12]) + (" …" if len(tags)>12 else "")).font = regular_font
            ws_dict.merge_cells(f"B{lr}:J{lr}")
            lr+=1
        ws_dict.freeze_panes = "A2"; ws_dict.auto_filter.ref = f"A1:J{r-1}" if r>2 else "A1:J1"
        for idx,w in enumerate([22,14,12,8,10,10,8,28,28,8], start=1):
            ws_dict.column_dimensions[get_column_letter(idx)].width = w

        # Print settings
        for ws in [ws_kpi, ws_data, ws_ref, ws_iss, ws_dup, ws_dict]:
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            ws.page_setup.orientation = "landscape"
            ws.page_setup.paperSize = ws.PAPERSIZE_A3

        wb.save(output_path_or_buffer)

